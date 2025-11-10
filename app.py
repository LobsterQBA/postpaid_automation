import pandas as pd
import streamlit as st
import re
from openpyxl import load_workbook
import io
from io import BytesIO

# ## Set configurations and title of the page
st.set_page_config(page_title="Postpaid Data Cleaning", layout="wide")
st.title("Postpaid Data Cleaning")

# How-To Section
st.subheader("Step-by-Step Guide")
with st.expander("📖 How to Use This App", expanded=False):
    st.markdown(""" 
    **1. Upload your files:** *You can clean all files at once or only upload one as you need, it is the same process*
    - In the **Email Data** section, upload:
        - Raw Email Data (PBI export) - High level open and click data by segment
        - Raw Email Click Data (AcV8 export) - CTA/URL clicks within the EM body
    - In the **SMS Data** section, identify if your data is from PBI or Branch.io. Then upload the file to the corresponding input:
        - Raw SMS Data (PBI export)
        - Raw SMS Data (Branch.io export) 
        - *Ensure the data source of the file matches the corresponding input on this page, or there will be an error*
    - In the **Deploy Doc** section, upload the campaign MD/DD excel workbook specific to the touch you are cleaning.

    **2. Campaign Details:** Use the sidebar on the left of the page to enter:
    - The campaign name you wish to be in the final clean data files 
    - Deploy date
    
    **3. Edit Delivery Label Values**  
    Adjust the table below directly on the page to match what values you need to pull from the delivery label, with which
    column you wish for them to go to.

    **4. Clean the Data**  
    Once files are uploaded, campaign details are inputted, and delivery label values are adjusted, click the 
    **Clean the data!** button to process your files.

    **5. Download Results**  
    Once cleaned, a preview of the clean data will appear. QA the data is what you expect and then download each dataset using the provided buttons.

    """)

# --------------------
# Email Section
# --------------------
st.subheader("Input the Raw Data File(s)")
with st.expander("📧 Email Data Inputs", expanded=False):
 raw_EM_uploaded_file = st.file_uploader(
     "Drop or Upload the Raw Email Data (PBI segment level, NOT CTA/URL clicks)", 
     type='xlsx', 
     key="raw_EM"
 )
 raw_EM_clicks_uploaded_file = st.file_uploader(
     "Drop or Upload the Raw Email from AcV8 (CTA/URL clicks)", 
     type='xlsx', 
     key="raw_EM_clicks"
 )

# --------------------
# SMS Section
# --------------------
with st.expander("💬 SMS Data Inputs", expanded=False):
 raw_SMSPBI_uploaded_file = st.file_uploader(
     "Drop or Upload the Raw SMS Data from **PBI**", 
     type="xlsx", 
     key="raw_SMSPBI"
 )
 raw_SMSbranch_uploaded_file = st.file_uploader(
     "Drop or Upload the Raw SMS Data from **Branch.io**", 
     type="xlsx", 
     key="raw_SMSbranch"
 )

# --------------------
# Deploy Doc Section
# --------------------
with st.expander("📄 Deploy Document (MD/DD)", expanded=False):
 DD_file = st.file_uploader(
     "Drop or Upload the DD", 
     type=["xlsx", "xlsm"], 
     key="DD_file"
 )

# --------------------
# Sidebar for campaign metadata
# --------------------
st.sidebar.header("📅 Campaign Details")
campaign_name_input = st.sidebar.text_input("Campaign Name")
deploy_date_input = st.sidebar.date_input("Deploy Date")


## Function to process raw EM file into a df
## show_raw = st.checkbox("Show uploaded raw EM data?")
if raw_EM_uploaded_file is not None:
    raw_EM_df = pd.read_excel(raw_EM_uploaded_file)
    st.write(raw_EM_df)

    ## Create new dataframe for the cleaned data with the correct column names
    ## Define column names
    columns = ["Campaign", "Deploy Date", "Delivery Label (Treatment)", "Touch", "OS", "Cohort", "SL Testing Variant", "Other Testing Variant", "Audience Details 1", "Audience Details 2", "Audience Details 3", "Subject Line", "Preheader", "Sends", "Deliveries", "Unique Opens", "Unique Clicks"]
    clean_EM_df = pd.DataFrame(columns=columns)

## Function to process raw EM clicks file into a df
## show_clicks = st.checkbox("Show uploaded EM clicks data?")
if raw_EM_clicks_uploaded_file is not None:
    raw_EM_clicks_df = pd.read_excel(raw_EM_clicks_uploaded_file)

    ## Create new dataframe for the cleaned data with the correct column names
    ## Define column names
    columns = ["Campaign", "Deploy Date", "Delivery Label (Treatment)", "Touch", "OS", "Cohort", "SL Testing Variant", "Other Testing Variant", "Audience Details 1", "Audience Details 2", "Audience Details 3", "CTA", "Position (Module #)", "CTA Offer Details", "CTA Category", "Link Style", "Device Category", "Device Type", "Deliveries", "Total Clicks", "CTR", "Click Share"]
    clean_EM_clicks_df = pd.DataFrame(columns=columns)

## Function to process raw SMS file into a df
## show_clicks = st.checkbox("Show uploaded SMS clicks data?")
if raw_SMSPBI_uploaded_file is not None:
    raw_SMS_df = pd.read_excel(raw_SMSPBI_uploaded_file)
    ## st.write(raw_SMS_df)

    ## Create new dataframe for the cleaned data with the correct column names
    ## Define column names
    columns = ["Campaign", "Deploy Date", "Delivery Label (Treatment)", "Touch", "OS", "Cohort", "SMS Testing Variant", "Other Testing Variant", "Audience Details 1", "Audience Details 2", "Audience Details 3", "Creative", "Sends", "Deliveries", "Unique Clicks", "CTR"]
    clean_SMS_df = pd.DataFrame(columns=columns)

## Define mapping: {new column : old column}
mapping = {
    "Deploy Date": "Deploy Date",
    "Delivery Label (Treatment)": ["Delivery Label", "DeliveryLabel", "Label"],
    "Sends": ["Sent", "Processed"],
    "Deliveries": ["Deliveries", "Received", "Success"],
    "Unique Opens": "Unique Opens",
    "Unique Clicks": "Unique Clicks",
    "CTA": "CTA",
    "Total Clicks": ["Clicks", "Total Clicks"]
}

# -------------------------
# Editable regex_map section
# -------------------------
regex_df = pd.DataFrame({
    "Column in Clean File": [
        "Touch", "OS", "Cohort", "SL Testing Variant", "SMS Testing Variant",
        "Other Testing Variant", "Audience Details 1", "Audience Details 2", "Audience Details 3"
    ],
    "Delivery Label Values": [
        r"(T1|T2|T3|T4|Preorder|Launch)",
        r"(IOS|AND)",
        r"(Growth|Churn)",
        r"(SLA|SLB|SLC)",
        r"(A|B|C)",
        r"(|)",
        r"(|)",
        r"(|)",
        r"(|)"
    ]
})

## Editable regex map directly on the page
st.subheader("Type in Values Needed from Delivery Label")
st.markdown("Ensure that what is in this mapping table is an exact match of the value you wish to pull from the delivery label (not case sensitive). Okay to leave unnecessary values blank.")
edited_regex_df = st.data_editor(regex_df, num_rows="dynamic")
regex_map = dict(zip(edited_regex_df["Column in Clean File"], edited_regex_df["Delivery Label Values"]))


## Function that performs EM data cleaning steps: Matching 1:1 columns and pulling values from the delivery label
def transform_data_EM(raw_EM_df, mapping, source_col, regex_map):
    ## Match columns from raw data to clean file
    for new_col, old_cols in mapping.items():
        # Ensure old_cols is always a list
        if not isinstance(old_cols, list):
            old_cols = [old_cols]

        # Find the first matching column in raw data
        found_col = next((col for col in old_cols if col in raw_EM_df.columns), None)

        if found_col is not None:
            col_data = raw_EM_df[found_col].fillna("")
            # Only round if the column is numeric
            if pd.api.types.is_numeric_dtype(col_data):
                col_data = col_data.round(0)
            clean_EM_df[new_col] = col_data

    ## Extract values from delivery label, leave blank if not found
    for new_col, pattern in regex_map.items():
        clean_EM_df[new_col] =clean_EM_df[source_col].str.extract(pattern, expand=False, flags=re.IGNORECASE).fillna("")

    ## Drop columns that are not specific to this EM data sheet brought in from the delivery label regex sheet (specifically the SMS Testing column)
    clean_EM_df.drop(columns=['SMS Testing Variant'], inplace=True)
    
    return clean_EM_df

## For the openpxyl search through the DD
## Normalize header text by removing special characters and extra spaces
def normalize_header(value):
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = text.replace("\xa0", " ").replace("\n", " ")
    text = re.sub(r"\s+", " ", text)
    return text

## Function to find subject lines and preheader
def find_SLs(clean_EM_df, DD_file, sheet_name_EM="EM_POST", delivery_col_name="Delivery Label", SL_col_name="Subject Line", preheader_col_name="Preheader", clean_EM_df_delivery_col="Delivery Label (Treatment)"):
    # Wrap the uploaded file for openpyxl
    DD_file.seek(0)
    in_mem_file = BytesIO(DD_file.read())
    wb = load_workbook(in_mem_file, data_only=True)
    ws = wb[sheet_name_EM]

    delivery_col_idx = None
    SL_col_idx = None
    preheader_col_idx = None

    ## Normalize header names for comparison using the predefined normalize_header function above
    delivery_col_name_norm = normalize_header(delivery_col_name)
    SL_col_name_norm = normalize_header(SL_col_name)
    preheader_col_name_norm = normalize_header(preheader_col_name)

    ## Find where the table actually starts in the DD and begin pulling info from there
    header_row = -1
    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        ## Normalize all cell values in this row
        normalized_values = [normalize_header(cell.value) for cell in row]

        ## Check if all three headers are present in this row (using startswith)
        has_delivery = any(val.startswith(delivery_col_name_norm) for val in normalized_values)
        has_subject  = any(val.startswith(SL_col_name_norm) for val in normalized_values)
        has_preheader = any(val.startswith(preheader_col_name_norm) for val in normalized_values)

        if has_delivery and has_subject and has_preheader:
        ## Assign indexes based on this row
            for cell in row:
                    cell_val_norm = normalize_header(cell.value)
                    if delivery_col_idx is None and cell_val_norm.startswith(delivery_col_name_norm):
                        delivery_col_idx = cell.column
                    elif SL_col_idx is None and cell_val_norm.startswith(SL_col_name_norm):
                        SL_col_idx = cell.column
                    elif preheader_col_idx is None and cell_val_norm.startswith(preheader_col_name_norm):
                        preheader_col_idx = cell.column
            if delivery_col_idx and SL_col_idx and preheader_col_idx:
                header_row = row_idx
                break

    ## Error handling if columns are not found
    if not (delivery_col_idx and SL_col_idx and preheader_col_idx):
        raise ValueError(f"Could not find '{delivery_col_name}' or '{SL_col_name_norm}' or '{preheader_col_idx}' in sheet '{sheet_name_EM}'")

    ## Store the subject lines and preheaders in a dictionary as strings, "decoded"
    subject_lookup = {}
    preheader_lookup = {}

    ## Iterate through rows after header
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        dl_value = ws.cell(row=row[0].row, column=delivery_col_idx).value
        sl_value = ws.cell(row=row[0].row, column=SL_col_idx).value
        ph_value = ws.cell(row=row[0].row, column=preheader_col_idx).value

        if dl_value is not None:
          norm_key = str(dl_value).strip().lower()

          # Only take the first occurrence for each delivery label
          if norm_key not in subject_lookup:
              subject_lookup[norm_key] = str(sl_value) if sl_value is not None else ""
          if norm_key not in preheader_lookup:
              preheader_lookup[norm_key] = str(ph_value) if ph_value is not None else ""

    # Map values back to DataFrame
    clean_EM_df["Subject Line"] = (
        clean_EM_df[clean_EM_df_delivery_col]
        .astype(str).str.strip().str.lower()
        .map(subject_lookup)
        .fillna("")
    )

    clean_EM_df["Preheader"] = (
        clean_EM_df[clean_EM_df_delivery_col]
        .astype(str).str.strip().str.lower()
        .map(preheader_lookup)
        .fillna("")
    )
    return clean_EM_df

## Function to perform EM clicks data cleaning steps: Matching 1:1 columns, pulling info from delivery label, calculating CTR & share of clicks
def transform_data_clicks(raw_EM_clicks_df, mapping, source_col, regex_map):
    ## Pull in metrics from raw data
    for new_col, old_cols in mapping.items():
        # Ensure old_cols is always a list
        if not isinstance(old_cols, list):
            old_cols = [old_cols]

        # Find the first matching column in raw data
        found_col = next((col for col in old_cols if col in raw_EM_clicks_df.columns), None)

        if found_col is not None:
            col_data = raw_EM_clicks_df[found_col].fillna("")
            # Only round if the column is numeric
            if pd.api.types.is_numeric_dtype(col_data):
                col_data = col_data.round(0)
            clean_EM_clicks_df[new_col] = col_data

    ## Pull out info from delivery label
    for new_col, pattern in regex_map.items():
        clean_EM_clicks_df[new_col] =clean_EM_clicks_df[source_col].str.extract(pattern, expand=False, flags=re.IGNORECASE).fillna("")

    ## Drop columns that are not specific to this EM data sheet brought in from the delivery label regex sheet (specifically the SMS Testing column)
    clean_EM_clicks_df.drop(columns=['Other Testing Variant'], inplace=True)

    ## Calculate CTR
    clean_EM_clicks_df["Deliveries"] = pd.to_numeric(clean_EM_clicks_df["Deliveries"], errors='coerce').fillna(0)
    clean_EM_clicks_df["Total Clicks"] = pd.to_numeric(clean_EM_clicks_df["Total Clicks"], errors='coerce').fillna(0)
    clean_EM_clicks_df["CTR"] = clean_EM_clicks_df.apply(
                lambda row: round(row["Total Clicks"] / row["Deliveries"], 5) if row["Deliveries"] != 0 else 0.00,
                axis=1
            )
    
    ## Calculate click share
    sum_clicks_per_label = {}
    sum_clicks_per_label = clean_EM_clicks_df.groupby('Delivery Label (Treatment)')['Total Clicks'].transform('sum')
    clean_EM_clicks_df["Click Share"] = clean_EM_clicks_df.apply(
                    lambda row: round(row["Total Clicks"] / sum_clicks_per_label[row.name], 5),
                    axis=1
                )
    
    return clean_EM_clicks_df

## Function that performs SMS data cleaning steps: Matching 1:1 columns, pulling values from the delivery label, calculate CTR
def transform_data_SMS(raw_SMS_df, mapping, source_col, regex_map):
    ## Match columns from raw data to clean file
    for new_col, old_cols in mapping.items():
        # Ensure old_cols is always a list
        if not isinstance(old_cols, list):
            old_cols = [old_cols]

        # Find the first matching column in raw data
        found_col = next((col for col in old_cols if col in raw_SMS_df.columns), None)

        if found_col is not None:
            col_data = raw_SMS_df[found_col].fillna("")
            # Only round if the column is numeric
            if pd.api.types.is_numeric_dtype(col_data):
                col_data = col_data.round(0)
            clean_SMS_df[new_col] = col_data

    ## Extract values from delivery label, leave blank if not found
    for new_col, pattern in regex_map.items():
        clean_SMS_df[new_col] =clean_SMS_df[source_col].str.extract(pattern, expand=False, flags=re.IGNORECASE).fillna("")

    ## Drop columns that are not specific to this EM data sheet brought in from the delivery label regex sheet (specifically the SMS Testing column)
    clean_SMS_df.drop(columns=['SL Testing Variant'], inplace=True)
    
    ## Calculate CTR
    clean_SMS_df["Deliveries"] = pd.to_numeric(clean_SMS_df["Deliveries"], errors='coerce').fillna(0)
    clean_SMS_df["Unique Clicks"] = pd.to_numeric(clean_SMS_df["Unique Clicks"], errors='coerce').fillna(0)
    clean_SMS_df["CTR"] = clean_SMS_df.apply(
                lambda row: round(row["Unique Clicks"] / row["Deliveries"], 5) if row["Deliveries"] != 0 else 0.00,
                axis=1
            )
    
    return clean_SMS_df

## Function to find SMS creative
def find_creative(clean_SMS_df, DD_file, sheet_name_SMS="SMS_POST", delivery_col_name="Delivery Label", creative_col_name="Creative Message", clean_SMS_df_delivery_col="Delivery Label (Treatment)"):
    # Wrap the uploaded file for openpyxl
    DD_file.seek(0)
    in_mem_file = BytesIO(DD_file.read())
    wb = load_workbook(in_mem_file, data_only=True)
    ws = wb[sheet_name_SMS]

    delivery_col_idx = None
    creative_col_idx = None

    ## Normalize header names for comparison
    delivery_col_name_norm = normalize_header(delivery_col_name)
    creative_col_name_norm = normalize_header(creative_col_name)

    ## Find where the table actually starts in the DD and begin pulling info from there
    header_row = -1
    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        ## Normalize all cell values in this row
        normalized_values = [normalize_header(cell.value) for cell in row]

        ## Check if all three headers are present in this row (using startswith)
        has_delivery = any(val.startswith(delivery_col_name_norm) for val in normalized_values)
        has_creative  = any(val.startswith(creative_col_name_norm) for val in normalized_values)

        if has_delivery and has_creative:
        ## Assign indexes based on this row
            for cell in row:
                    cell_val_norm = normalize_header(cell.value)
                    if delivery_col_idx is None and cell_val_norm.startswith(delivery_col_name_norm):
                        delivery_col_idx = cell.column
                    elif creative_col_idx is None and cell_val_norm.startswith(creative_col_name_norm):
                        creative_col_idx = cell.column
            if delivery_col_idx and creative_col_idx:
                header_row = row_idx
                break

    if not delivery_col_idx or not creative_col_idx:
        raise ValueError(f"Could not find '{delivery_col_name}' or '{creative_col_name}' in sheet '{sheet_name_SMS}'")

    ## Store the creatives in a dictionary as strings, "decoded"
    excel_lookup = {}
    for row in ws.iter_rows(min_row=header_row+1, max_row=ws.max_row):
        dl_value = ws.cell(row=row[0].row, column=delivery_col_idx).value
        creative_value = ws.cell(row=row[0].row, column=creative_col_idx).value
        if dl_value is not None:
            norm_key = str(dl_value).strip().lower()
            if creative_value is not None:
                try:
                    ## Fix smart quotes/encoding issues
                    creative_value = str(creative_value).encode("latin1").decode("utf-8")
                except Exception:
                    creative_value = str(creative_value)
            excel_lookup[norm_key] = creative_value

    ## Normalize the DataFrame column before mapping to avoid errors
    clean_SMS_df["Creative"] = (
        clean_SMS_df[clean_SMS_df_delivery_col]
        .astype(str).str.strip().str.lower()
        .map(excel_lookup)
        .fillna("")
    )

    return clean_SMS_df

# Button to run the data cleaning function
st.subheader("Clean the data")
if st.button("Clean the data!"):
    
    # EM data cleaning
    if raw_EM_uploaded_file is not None:
        clean_EM_df = transform_data_EM(raw_EM_df, mapping, "Delivery Label (Treatment)", regex_map)
        if DD_file is not None:
            final_EM_df = find_SLs(clean_EM_df, DD_file, sheet_name_EM="EM_POST")
        else:
            final_EM_df = clean_EM_df.copy()
        
        # Apply campaign metadata
        if campaign_name_input:
            final_EM_df["Campaign"] = campaign_name_input
        if deploy_date_input:
            final_EM_df["Deploy Date"] = pd.to_datetime(deploy_date_input)

        st.session_state["final_EM_df"] = final_EM_df
        
    
    # EM clicks cleaning
    if raw_EM_clicks_uploaded_file is not None:
        final_EM_clicks_df = transform_data_clicks(raw_EM_clicks_df, mapping, "Delivery Label (Treatment)", regex_map)

        # Apply campaign metadata
        if campaign_name_input:
            final_EM_clicks_df["Campaign"] = campaign_name_input
        if deploy_date_input:
            final_EM_clicks_df["Deploy Date"] = pd.to_datetime(deploy_date_input)
        
        # # Show preview
        # st.subheader("Cleaned EM Clicks Data")
        # st.write(final_EM_clicks_df)

        st.session_state["final_EM_clicks_df"] = final_EM_clicks_df


    # SMS data cleaning
    if raw_SMSPBI_uploaded_file is not None:
        clean_SMS_df = transform_data_SMS(raw_SMS_df, mapping, "Delivery Label (Treatment)", regex_map)
        if DD_file is not None:
            final_SMS_df = find_creative(clean_SMS_df, DD_file, sheet_name_SMS="SMS_POST")
        else:
            final_SMS_df = clean_EM_df.copy()
        
        # Apply campaign metadata
        if campaign_name_input:
            final_SMS_df["Campaign"] = campaign_name_input
        if deploy_date_input:
            final_SMS_df["Deploy Date"] = pd.to_datetime(deploy_date_input)

        st.session_state["final_SMS_df"] = final_SMS_df
        


# Show download buttons if data exists in session_state
if "final_EM_df" in st.session_state:
    st.subheader("Cleaned EM Data")
    st.write(st.session_state["final_EM_df"])
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
        st.session_state["final_EM_df"].to_excel(writer, sheet_name='EM', index=False)
    buffer.seek(0)
    st.download_button(
        label="Download Clean EM Data",
        data=buffer,
        file_name="clean_EM.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

if "final_EM_clicks_df" in st.session_state:
    st.subheader("Cleaned EM Clicks Data")
    st.write(st.session_state["final_EM_clicks_df"])
    buffer_clicks = io.BytesIO()
    with pd.ExcelWriter(buffer_clicks, engine='xlsxwriter') as writer:
        st.session_state["final_EM_clicks_df"].to_excel(writer, sheet_name='Clicks', index=False)
    buffer_clicks.seek(0)
    st.download_button(
        label="Download Clean EM Clicks Data",
        data=buffer_clicks,
        file_name="clean_EM_clicks.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

if "final_SMS_df" in st.session_state:
    st.subheader("Cleaned SMS Data")
    st.write(st.session_state["final_SMS_df"])
    buffer_sms = io.BytesIO()
    with pd.ExcelWriter(buffer_sms, engine='xlsxwriter') as writer:
        st.session_state["final_SMS_df"].to_excel(writer, sheet_name='SMS', index=False)
    buffer_sms.seek(0)
    st.download_button(
        label="Download Clean SMS Data",
        data=buffer_sms,
        file_name="clean_SMS.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
