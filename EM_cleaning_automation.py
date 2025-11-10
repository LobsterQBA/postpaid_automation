import pandas as pd
import re
from openpyxl import load_workbook
from io import BytesIO

## Read in raw EM data file
raw_EM_df = pd.read_excel("PBI_raw_EM.xlsx")

## Pull in SLs and preheader from DD
file_path = "CD_4233611_8620110_Back_To_School_EM-SMS_T2.xlsm"
sheet_name = "EM_POST"

## Create new dataframe for the cleaned data with the correct column names
## Define column names
columns = ["Campaign", "Deploy Date", "Delivery Label (Treatment)", "Touch", "OS", "Cohort", "Other Audience Details", "Sends", "Deliveries", "Unique Opens", "Unique Clicks", "Testing Variant", "Subject Line", "Preheader"]
clean_EM_df = pd.DataFrame(columns=columns)

## Define mapping: {new column : old column}
mapping = {
    "Deploy Date": "Deploy Date",
    "Delivery Label (Treatment)": ["Delivery Label", "DeliveryLabel", "Label"],
    "Sends": ["Sent", "Processed"],
    "Deliveries": ["Deliveries", "Received", "Success"],
    "Unique Opens": "Unique Opens",
    "Unique Clicks": "Unique Clicks",
    "CTA": "CTA",
    "Total Clicks": ["Clicks", "Total Clicks"]
}

# Define all regex patterns
regex_map = {
    "Touch": r"(T1|T2|T3|T4)",
    "OS": r"(IOS|AND)",
    "Cohort": r"(Growth|Churn)",
    "Other Audience Details": r"(k12|college)",
    "Testing Variant": r"(SLA|SLB)"
}

## Combine the two functions into one
def transform_data(raw_EM_df, mapping, source_col, regex_map):
    ## Match columns from raw data to clean file
    for new_col, old_cols in mapping.items():
        # Ensure old_cols is always a list
        if not isinstance(old_cols, list):
            old_cols = [old_cols]

        # Find the first matching column in raw data
        found_col = next((col for col in old_cols if col in raw_EM_df.columns), None)

        if found_col is not None:
            col_data = raw_EM_df[found_col].fillna("")
            # Only round if the column is numeric
            if pd.api.types.is_numeric_dtype(col_data):
                col_data = col_data.round(0)
            clean_EM_df[new_col] = col_data

    ## Pull values from delivery label
    for new_col, pattern in regex_map.items():
        clean_EM_df[new_col] =clean_EM_df[source_col].str.extract(pattern, expand=False, flags=re.IGNORECASE).fillna("")
    
    return clean_EM_df

## For the openpxyl search through the DD
## Normalize header text by removing special characters and extra spaces
def normalize_header(value):
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = text.replace("\xa0", " ").replace("\n", " ")
    text = re.sub(r"\s+", " ", text)
    return text

## Function to find subject lines and preheader
def find_SLs(clean_EM_df, DD_file, sheet_name_EM="EM_POST", delivery_col_name="Delivery Label", SL_col_name="Subject Line", preheader_col_name="Preheader", clean_EM_df_delivery_col="Delivery Label (Treatment)"):
    # Wrap the uploaded file for openpyxl
    ## DD_file.seek(0)
    ## in_mem_file = BytesIO(DD_file.read())
    wb = load_workbook(DD_file, data_only=True)
    ws = wb[sheet_name_EM]

    delivery_col_idx = None
    SL_col_idx = None
    preheader_col_idx = None

    ## Normalize header names for comparison using the predefined normalize_header function above
    delivery_col_name_norm = normalize_header(delivery_col_name)
    SL_col_name_norm = normalize_header(SL_col_name)
    preheader_col_name_norm = normalize_header(preheader_col_name)

    ## Find where the table actually starts in the DD and begin pulling info from there
    header_row = -1
    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        ## Normalize all cell values in this row
        normalized_values = [normalize_header(cell.value) for cell in row]

        ## Check if all three headers are present in this row (using startswith)
        has_delivery = any(val.startswith(delivery_col_name_norm) for val in normalized_values)
        has_subject  = any(val.startswith(SL_col_name_norm) for val in normalized_values)
        has_preheader = any(val.startswith(preheader_col_name_norm) for val in normalized_values)

        if has_delivery and has_subject and has_preheader:
        ## Assign indexes based on this row
            for cell in row:
                    cell_val_norm = normalize_header(cell.value)
                    if delivery_col_idx is None and cell_val_norm.startswith(delivery_col_name_norm):
                        delivery_col_idx = cell.column
                    elif SL_col_idx is None and cell_val_norm.startswith(SL_col_name_norm):
                        SL_col_idx = cell.column
                    elif preheader_col_idx is None and cell_val_norm.startswith(preheader_col_name_norm):
                        preheader_col_idx = cell.column
            if delivery_col_idx and SL_col_idx and preheader_col_idx:
                header_row = row_idx
                break

    ## Error handling if columns are not found
    if not (delivery_col_idx and SL_col_idx and preheader_col_idx):
        raise ValueError(f"Could not find '{delivery_col_name}' or '{SL_col_name_norm}' or '{preheader_col_idx}' in sheet '{sheet_name_EM}'")

    ## Store the subject lines and preheaders in a dictionary as strings, "decoded"
    subject_lookup = {}
    preheader_lookup = {}

    ## Iterate through rows after header
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        dl_value = ws.cell(row=row[0].row, column=delivery_col_idx).value
        sl_value = ws.cell(row=row[0].row, column=SL_col_idx).value
        ph_value = ws.cell(row=row[0].row, column=preheader_col_idx).value

        if dl_value is not None:
          norm_key = str(dl_value).strip().lower()

          # Only take the first occurrence for each delivery label
          if norm_key not in subject_lookup:
              subject_lookup[norm_key] = str(sl_value) if sl_value is not None else ""
          if norm_key not in preheader_lookup:
              preheader_lookup[norm_key] = str(ph_value) if ph_value is not None else ""

    # Map values back to DataFrame
    clean_EM_df["Subject Line"] = (
        clean_EM_df[clean_EM_df_delivery_col]
        .astype(str).str.strip().str.lower()
        .map(subject_lookup)
        .fillna("")
    )

    clean_EM_df["Preheader"] = (
        clean_EM_df[clean_EM_df_delivery_col]
        .astype(str).str.strip().str.lower()
        .map(preheader_lookup)
        .fillna("")
    )
    return clean_EM_df

## Run the functions
clean_EM_df = transform_data(raw_EM_df, mapping, "Delivery Label (Treatment)", regex_map)
final_EM_df = find_SLs(clean_EM_df, file_path, sheet_name)

## Save the clean EM data as a csv file
final_EM_df.to_excel('clean_EM.xlsx', index=False)

